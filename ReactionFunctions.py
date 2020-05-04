import openpyxl
from random import randint
from os import listdir

workFile = openpyxl.load_workbook("Stats.xlsx", data_only=True)
document = workFile.active


def calculateWinLoss(x, y):
    if y == 0 or y == "0":
        return "No losses"
    return int(x) / int(y)


def winOrLoss(me, split_command):
    workFile = openpyxl.load_workbook("Stats.xlsx", data_only=True)
    document = workFile.active
    # find discord ID in spreadsheet
    # go to win column, get the data and add plus one
    # go to the other person mentioned and go to their loss sheet and add one to their loss sheet
    # time to define args
    # me is the sender
    Foundhim = False
    Foundme = False
    max_height = document.max_row
    max_width = document.max_column
    rowMe = 0
    colMe = 0
    rowHim = 0
    colHim = 0
    try:
        for char in split_command[3]:
            if char in "<>@!":
                split_command[3] = split_command[3].replace(char, '')
    except IndexError:
        return "You gave an incorrect input, try again."
    for i in range(2, max_height + 1):
        if Foundme == False and str(document.cell(row=i, column=2).value) == str(me):
            rowMe = i
            colMe = 1
            Foundme = True
        if Foundhim == False and str(document.cell(row=i, column=2).value) == str(split_command[3]):
            rowHim = i
            colHim = 1
            Foundhim = True
    if Foundhim and Foundme:
        print("Both people were found")
        # This is where the updating happened
        if (split_command[1].lower() == "win"):
            print("updating in win")
            document.cell(row=rowMe, column=colMe + 2).value = str(
                int(document.cell(row=rowMe, column=colMe + 2).value) + 1)
            document.cell(row=rowHim, column=colHim + 3).value = str(
                int(document.cell(row=rowHim, column=colMe + 3)) + 1)
            document.cell(row=rowMe, column=5).value = calculateWinLoss(document.cell(row=rowMe, column=3).value,
                                                                        document.cell(row=rowMe, column=4).value)
            document.cell(row=rowMe, column=5).value = calculateWinLoss(document.cell(row=rowHim, column=3).value,
                                                                        document.cell(row=rowHim, column=4).value)
            workFile.save("Stats.xlsx")
            return "Data has been updated, please do !falcon record to see your new record"
        elif split_command[1].lower() == "loss":
            print("updating in loss")
            document.cell(row=rowMe, column=colMe + 3).value = str(
                int(document.cell(row=rowMe, column=colMe + 3).value) + 1)
            document.cell(row=rowHim, column=colHim + 2).value = str(
                int(document.cell(row=rowHim, column=colMe + 2).value) + 1)
            document.cell(row=rowMe, column=5).value = calculateWinLoss(document.cell(row=rowMe, column=3).value,
                                                                        document.cell(row=rowMe, column=4).value)
            document.cell(row=rowMe, column=5).value = calculateWinLoss(document.cell(row=rowHim, column=3).value,
                                                                        document.cell(row=rowHim, column=4).value)
            workFile.save("Stats.xlsx")
            return "Data has been updated, please do !falcon record to see your new record"
        else:
            print("Something is wrong")
            return False
    else:
        return "One of you 2 is not in the league, please make sure you both appear"


# Will sort through all the players and print them
# in order of highest to lowest Win/Loss Ratio
def rankList(args):
    stat_list = {}
    max_height = document.max_row
    for i in range(2, max_height + 1):
        if (document.cell(row=i, column=1).value != "None"):
            stat_list[document.cell(row=i, column=1).value] = str(document.cell(row=i, column=5).value)
    # We will sort the stat list here
    a = sorted(stat_list, key=stat_list.get)
    a.reverse()
    print(a)
    pos = 1
    final_str = "Rankings:\n"
    for i in a:
        final_str += str(pos) + " -> " + i + " : " + stat_list[i] + "\n"
        # When we find your position, we will append that message to the beginning of the list and say something
        # of the like, we found you at position number x.
        if (i == args[0]):
            final_str = "You were found at position number " + str(pos) + "\n" + final_str
        pos += 1
    return final_str


def displayRecord(args):
    workFile = openpyxl.load_workbook("Stats.xlsx", data_only=True)
    document = workFile.active
    # Will display the record of the calling person
    # Will go in by his ID and print out his
    # win loss record as well as his W/L ratio
    str_result = ""
    isThere = False
    row = 0
    col = 0
    max_height = document.max_row
    for i in range(2, max_height + 1):

        if (args[0] == document.cell(row=i, column=1).value):
            row = i
            col = 1
            isThere = True
            break
    for i in range(2, max_height + 1):
        if (args[0] == document.cell(row=i, column=2).value):
            row = i
            col = 2
            isThere = True
            break
    if isThere:
        if col == 1:
            str_result = "The record of " + str(document.cell(row=row, column=col).value) + " is " + str(
                document.cell(row=row, column=3).value) + " wins and " + str(
                document.cell(row=row, column=4).value) + " losses for a Win-Loss ratio of " + str(
                document.cell(row=row, column=5).value)
        elif col == 2:
            str_result = "The record of " + str(document.cell(row, col - 1).value) + " is " + str(
                document.cell(row, 3).value) + " wins and " + str(
                document.cell(row, 4).value) + " losses for a Win-Loss ratio of " + str(document.cell(row, 5).value)
        return str_result
    else:
        return "You are not in the league"


def joinLeague(args):
    workFile = openpyxl.load_workbook("Stats.xlsx", data_only=True)
    document = workFile.active
    row_where_to_add = document.max_row + 1
    max_height = document.max_row
    for i in range(2, max_height + 1):
        if (document.cell(row=i, column=1).value == args[0]):
            return "You are already in the league, you cannot join again"

    document.cell(row=row_where_to_add, column=1).value = args[0]
    document.cell(row=row_where_to_add, column=2).value = str(args[1])
    document.cell(row=row_where_to_add, column=3).value = 0
    document.cell(row=row_where_to_add, column=4).value = 0
    document.cell(row=row_where_to_add, column=5).value = "No losses"
    workFile.save("Stats.xlsx")
    return args[0] + " has joined the league"


def listLeagueMembers():
    workFile = openpyxl.load_workbook("Stats.xlsx", data_only=True)
    document = workFile.active
    list = ""
    for i in range(2, document.max_row + 1):
        if (str(document.cell(row=i, column=1).value) != "None"):
            list += str(document.cell(row=i, column=1).value) + "\n"
    return "The current list of members is :\n" + list

def returnFunnyLink():
    from random import randint
    lmao=open("funnyLinks.txt","r")
    links=lmao.readlines()
    return links[randint(0, len(links) - 1)]

def listCommands():
    return "Here are the following commands:\n!falcon followed by\n1- win\loss vs @opponent\n2-join\n3-rank\n4-record\n5-other secret commands have fun finding them"

def handHolding():
    print ("Opening file")
    file = open("HandHolding.txt","r")
    gif = file.readlines()
    file.close()
    print ("Closing file")
    return gif[randint(0,len(gif)-1)]

def randomRacistReplies():
    print("Opening file")
    file = open("MR/racism.txt", "r")
    comments = file.readlines()
    print(comments)
    file.close()
    print("Closing file")
    return comments[randint(0, len(comments) - 1)]


def pullPictures(directory):
    pictures = listdir(directory)
    a = pictures[randint(0, len(pictures) - 1)]
    print(a)
    return a


def wholesomeReplies():
    print("Opening directories")
    return pullPictures("Wholesome")
